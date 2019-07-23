import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import os
import time
from tkinter import filedialog, Tk
from tkinter.filedialog import askdirectory

class Finder:
    def __init__(self):
        root = Tk()
        root.withdraw()
        # inputs
        print("Choose folder with csv files")
        self.input_folder = askdirectory()

        os.system("cls")
        print("Choose excel file whitch will be modified")
        self.excel =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))

        os.system("cls")
        print("Save excel file as:")
        self.save = filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))

        os.system("cls")
        print("I am working please take a coffee break")
        # excel
        self.book = openpyxl.load_workbook(self.excel)
        self.sheet = self.book.active

        self.csv_numbers = []
        self.serial_numbers = []
        self.serial_numbers_in_csv = []

        self.redFill = PatternFill(start_color="00b0f0",
                              end_color="00b0f0",
                              fill_type='solid')

    def get_serial_numbers(self):
        for i in range(len(self.sheet["Q"]) - 1):
            serial_number = str(self.sheet.cell(row=i+2, column=17).value)
            self.serial_numbers.append(serial_number.upper().replace("Z", "Y"))
        return self.serial_numbers


    def get_csv_numbers(self):
        for f in os.listdir(self.input_folder):
            path = self.input_folder + "\\" + f
            with open(path, newline='') as inputfile:
                for row in csv.reader(inputfile):
                    self.csv_numbers.append(row[0])
        return self.csv_numbers

    def finder(self):
        counter = 0
        for i in range(len(self.serial_numbers)):
            if self.serial_numbers[i] in self.csv_numbers:
                counter += 1
                self.serial_numbers_in_csv.append(self.serial_numbers[i])


    def finder_test(self):
        counter = 0
        for i in range(len(self.serial_numbers)):
            for j in range(len(self.csv_numbers)):
                if self.serial_numbers[i] == self.csv_numbers[j]:
                    counter += 1
                    break  # bez duplicit
            print(counter)

    def coloring(self):
        for i in range(len(self.sheet["Q"])):
            cell = str(self.sheet.cell(row=i + 1, column=17).value).upper().replace("Z", "Y")
            if cell in self.serial_numbers_in_csv:
                self.sheet.cell(row=i + 1, column=17).fill = self.redFill
        self.book.save(self.save)



start = time.time()

x = Finder()
x.get_serial_numbers()
x.get_csv_numbers()
x.finder()
x.coloring()

print("Done!")

print((time.time() - start) / 60)
